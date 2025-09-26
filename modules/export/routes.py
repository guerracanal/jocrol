
import io
import os
from flask import Blueprint, request, redirect, url_for, flash, send_file
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from data.data_manager import (
    JUEGOS_COLECCIONES_COLLECTION,
    CLIENTES_COLLECTION,
    EVENTOS_COLLECTION,
    LANZAMIENTOS_COLLECTION,
    RESERVAS_COLLECTION
)
from modules.clientes.services import obtener_clientes_todos, crear_cliente
from modules.eventos.services import obtener_eventos_todos, crear_evento, obtener_juegos_y_colecciones
from modules.lanzamientos.services import obtener_lanzamientos_todos, crear_lanzamiento
from modules.reservas.services import obtener_reservas_todas, crear_reserva

export_bp = Blueprint('export', __name__, url_prefix='/export')

def write_data_to_sheet(wb, sheet_name, data, headers_to_exclude=None):
    """Crea una hoja y escribe datos en ella."""
    ws = wb.create_sheet(title=sheet_name)
    if not data:
        ws.cell(row=1, column=1, value="No hay datos disponibles.")
        return

    headers = [key for key in data[0].keys() if key not in (headers_to_exclude or [])]
    ws.append(headers)

    for item in data:
        row = [item.get(h, '') for h in headers]
        ws.append(row)

@export_bp.route('/excel')
def export_excel():
    """Exporta todos los datos a un único fichero Excel con múltiples hojas."""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 1. Cargar todos los datos desde los servicios
    juegos_data = obtener_juegos_y_colecciones()
    clientes_data = obtener_clientes_todos()
    eventos_data = obtener_eventos_todos()
    lanzamientos_data = obtener_lanzamientos_todos()
    reservas_data = obtener_reservas_todas()

    # Mapeos para búsquedas rápidas en la hoja de Reservas
    clientes_map = {c['id']: c for c in clientes_data}
    lanzamientos_map = {l['id']: l for l in lanzamientos_data}
    eventos_map = {e['id']: e for e in eventos_data}
    
    # 2. Hoja de Juegos y Colecciones
    ws_juegos = wb.create_sheet(title='Juegos')
    ws_juegos.append(['Juego', 'Color'])
    ws_colecciones = wb.create_sheet(title='Colecciones')
    ws_colecciones.append(['Coleccion', 'Juego'])

    if isinstance(juegos_data, dict) and 'juegos' in juegos_data:
        for juego, details in juegos_data.get('juegos', {}).items():
            ws_juegos.append([juego, details.get('color', '')])
            for coleccion in details.get('colecciones', []):
                ws_colecciones.append([coleccion, juego])

    # 3. Hojas de Clientes, Eventos, Lanzamientos
    write_data_to_sheet(wb, 'Clientes', clientes_data, headers_to_exclude=['id'])
    write_data_to_sheet(wb, 'Eventos', eventos_data, headers_to_exclude=['id'])
    write_data_to_sheet(wb, 'Lanzamientos', lanzamientos_data, headers_to_exclude=['id'])

    # 4. Hoja de Reservas (con tratamiento especial)
    ws_reservas = wb.create_sheet(title='Reservas')
    reservas_headers = ['telefono_cliente', 'tipo', 'nombre', 'juego', 'coleccion', 'cantidad', 'fecha_reserva', 'estado', 'pagado', 'tipo_pago', 'notas']
    ws_reservas.append(reservas_headers)

    for reserva in reservas_data:
        cliente_telefono = clientes_map.get(reserva.get('cliente_id'), {}).get('telefono', '-')
        
        tipo, nombre, juego, coleccion = '', '', '', ''
        item = None

        if reserva.get('lanzamiento_id'):
            item = lanzamientos_map.get(reserva.get('lanzamiento_id'))
        elif reserva.get('evento_id'):
            item = eventos_map.get(reserva.get('evento_id'))

        if item:
            tipo = item.get('tipo', 'Lanzamiento') # Asumir 'Lanzamiento' si no hay tipo
            nombre = item.get('nombre', '')
            juego = item.get('juego', '')
            coleccion = item.get('coleccion', '')

        row = [
            cliente_telefono,
            tipo,
            nombre,
            juego,
            coleccion,
            reserva.get('cantidad', ''),
            reserva.get('fecha_reserva', ''),
            reserva.get('estado', ''),
            reserva.get('pagado', ''),
            reserva.get('tipo_pago', ''),
            reserva.get('notas', '')
        ]
        ws_reservas.append(row)

    # 5. Guardar y enviar el fichero
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    return send_file(
        excel_stream,
        as_attachment=True,
        download_name='export_gestion_lanzamientos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@export_bp.route('/import/excel', methods=['POST'])
def import_excel():
    if 'excel_file' not in request.files or not request.files['excel_file'].filename:
        flash('No se ha seleccionado ningún fichero.', 'danger')
        return redirect(url_for('main.index'))

    file = request.files['excel_file']
    sheet_to_import = request.form.get('sheet_name')
    import_mode = request.form.get('import_mode', 'overwrite')

    if not file.filename.endswith('.xlsx'):
        flash('Formato de fichero no válido. Por favor, sube un fichero .xlsx.', 'danger')
        return redirect(url_for('main.index'))
    if not sheet_to_import:
        flash('No se ha especificado qué hoja importar.', 'danger')
        return redirect(url_for('main.index'))

    try:
        wb = load_workbook(file)
        if sheet_to_import not in wb.sheetnames:
            flash(f"La hoja '{sheet_to_import}' no existe en el fichero.", 'danger')
            return redirect(url_for('main.index'))

        ws = wb[sheet_to_import]
        headers = [cell.value for cell in ws[1]]
        new_data_rows = [dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]

        # Lógica para 'Juegos' y 'Colecciones'
        if sheet_to_import in ['Juegos', 'Colecciones']:
            data = JUEGOS_COLECCIONES_COLLECTION.find_one({}, {'_id': 0}) or {'juegos': {}}
            
            if import_mode == 'overwrite':
                if sheet_to_import == 'Juegos':
                    data['juegos'] = {item.get('Juego'): {'color': item.get('Color', ''), 'colecciones': []} for item in new_data_rows}
                else: # Colecciones
                    flash('Para sobrescribir colecciones, por favor, sobrescriba la hoja "Juegos" y luego añada las colecciones.', 'warning')
                    return redirect(url_for('main.index'))
            else: # Append
                if sheet_to_import == 'Juegos':
                    for item in new_data_rows:
                        juego = item.get('Juego')
                        if not juego: continue
                        if juego not in data['juegos']:
                            data['juegos'][juego] = {'color': item.get('Color', ''), 'colecciones': []}
                        else:
                            data['juegos'][juego]['color'] = item.get('Color', data['juegos'][juego].get('color', ''))
                elif sheet_to_import == 'Colecciones':
                    for item in new_data_rows:
                        juego, coleccion = item.get('Juego'), item.get('Coleccion')
                        if not juego or not coleccion: continue
                        if juego in data['juegos'] and coleccion not in data['juegos'][juego]['colecciones']:
                            data['juegos'][juego]['colecciones'].append(coleccion)
            
            JUEGOS_COLECCIONES_COLLECTION.delete_many({})
            JUEGOS_COLECCIONES_COLLECTION.insert_one(data)

        # Lógica para hojas estándar
        else:
            collection_map = {
                'Clientes': (CLIENTES_COLLECTION, crear_cliente),
                'Eventos': (EVENTOS_COLLECTION, crear_evento),
                'Lanzamientos': (LANZAMIENTOS_COLLECTION, crear_lanzamiento)
            }
            
            if sheet_to_import in collection_map:
                collection, create_fn = collection_map[sheet_to_import]
                if import_mode == 'overwrite':
                    collection.delete_many({})
                
                for item_data in new_data_rows:
                    # Filtrar claves con valores None para no pasarlas a la función de creación
                    filtered_data = {k: v for k, v in item_data.items() if v is not None}
                    create_fn(**filtered_data)

            elif sheet_to_import == 'Reservas':
                if import_mode == 'overwrite':
                    RESERVAS_COLLECTION.delete_many({})
                
                clientes_map = {c['telefono']: c['id'] for c in obtener_clientes_todos()}
                lanzamientos_map = {(l['nombre'], l.get('juego', ''), l.get('coleccion', '')): l['id'] for l in obtener_lanzamientos_todos()}
                eventos_map = {(e['nombre'], e.get('juego', ''), e.get('coleccion', '')): e['id'] for e in obtener_eventos_todos()}

                for item in new_data_rows:
                    cliente_id = clientes_map.get(str(item.get('telefono_cliente')))
                    if not cliente_id:
                        flash(f"Cliente con teléfono {item.get('telefono_cliente')} no encontrado. Saltando reserva.", 'warning')
                        continue

                    lanzamiento_id, evento_id = None, None
                    item_key = (item.get('nombre'), item.get('juego', ''), item.get('coleccion', ''))
                    
                    if item.get('tipo') == 'Evento':
                        evento_id = eventos_map.get(item_key)
                    else: # Asumir Lanzamiento por defecto
                        lanzamiento_id = lanzamientos_map.get(item_key)

                    if not lanzamiento_id and not evento_id:
                        flash(f"Producto no encontrado para la reserva: {item_key}. Saltando.", 'warning')
                        continue
                    
                    reserva_data = {
                        'cliente_id': cliente_id,
                        'lanzamiento_id': lanzamiento_id,
                        'evento_id': evento_id,
                        'cantidad': item.get('cantidad'),
                        'fecha_reserva': item.get('fecha_reserva'),
                        'estado': item.get('estado'),
                        'pagado': item.get('pagado'),
                        'tipo_pago': item.get('tipo_pago'),
                        'notas': item.get('notas')
                    }
                    crear_reserva(**{k: v for k, v in reserva_data.items() if v is not None})
            else:
                flash(f"La importación para la hoja '{sheet_to_import}' no está implementada.", 'danger')
                return redirect(url_for('main.index'))

        flash(f'Datos de la hoja "{sheet_to_import}" importados correctamente en modo "{import_mode}".', 'success')

    except Exception as e:
        flash(f'Error al procesar el fichero Excel: {e}', 'danger')

    return redirect(url_for('main.index'))
